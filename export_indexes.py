"""
Export all records from both AI Search indexes to an Excel file.
Each index gets its own sheet.
"""
import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
import os, json

load_dotenv(override=True)

ENDPOINT = os.environ["AZURE_SEARCH_ENDPOINT"].rstrip("/")
API_KEY = os.environ["AZURE_SEARCH_API_KEY"]
API_VERSION = os.getenv("AZURE_SEARCH_API_VERSION", "2024-07-01")
GROUP_INDEX = os.getenv("AZURE_SEARCH_GROUP_INDEX", "group-slot-mapping-index")
USER_INDEX = os.getenv("AZURE_SEARCH_USER_INDEX", "user-slot-mapping-index")

HEADERS = {"api-key": API_KEY, "Content-Type": "application/json"}


def fetch_all(index: str, select: str) -> list[dict]:
    """Fetch all documents from an index using search with wildcard '*'."""
    url = f"{ENDPOINT}/indexes/{index}/docs/search?api-version={API_VERSION}"
    results = []
    body = {"search": "*", "select": select, "top": 200, "count": True}

    with httpx.Client(timeout=30) as client:
        while True:
            resp = client.post(url, headers=HEADERS, json=body)
            resp.raise_for_status()
            data = resp.json()
            docs = data.get("value", [])
            results.extend(docs)
            print(f"  Fetched {len(results)} records so far from {index}...", flush=True)

            # Check for continuation token
            next_link = data.get("@odata.nextLink") or data.get("@search.nextPageParameters")
            if next_link and isinstance(next_link, str):
                # nextLink is a full URL
                resp = client.get(next_link, headers=HEADERS)
                resp.raise_for_status()
                data = resp.json()
                docs = data.get("value", [])
                if not docs:
                    break
                results.extend(docs)
            elif next_link and isinstance(next_link, dict):
                body = next_link
                body["select"] = select
            else:
                break

    return results


def write_sheet(wb: openpyxl.Workbook, title: str, records: list[dict], columns: list[str]):
    """Write records to a named worksheet with formatting."""
    ws = wb.create_sheet(title=title)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for row_idx, rec in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = rec.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(records) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 50)

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    print(f"  Wrote {len(records)} rows to sheet '{title}'", flush=True)


def main():
    print("=" * 50, flush=True)
    print("  Exporting AI Search indexes to Excel", flush=True)
    print("=" * 50, flush=True)

    # Fetch group data
    print(f"\nFetching from {GROUP_INDEX}...", flush=True)
    group_cols = ["GroupID", "GroupName", "AlternateName1", "AlternateName2", "AlternateName3"]
    groups = fetch_all(GROUP_INDEX, ",".join(group_cols))

    # Fetch user data
    print(f"\nFetching from {USER_INDEX}...", flush=True)
    user_cols = ["id", "FirstName", "LastName"]
    users = fetch_all(USER_INDEX, ",".join(user_cols))

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_sheet(wb, "Groups", groups, group_cols)
    write_sheet(wb, "Users", users, user_cols)

    out_file = "SIRE_AI_Search_Data.xlsx"
    wb.save(out_file)
    print(f"\n✓ Saved to {out_file}", flush=True)
    print(f"  Groups: {len(groups)} records", flush=True)
    print(f"  Users:  {len(users)} records", flush=True)


if __name__ == "__main__":
    main()
